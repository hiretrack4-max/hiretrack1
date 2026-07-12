"""
Module 8 — Recruitment report generation (Excel / CSV / PDF).

Design
------
* **Grain** — one row per :class:`~core.models.CandidateJobMapping` (candidate ↔
  job). The BRD report columns mix Job fields (Role / Status), Candidate fields
  (name, experience, salary, notice, status), the Recruiter (on the mapping), the
  latest Interview Date (on the mapping) and the Offer Status (on the mapping), so
  the mapping is the natural grain that carries all of them.

* **Efficient queries** — the whole report is one query. ``select_related``
  pulls Candidate / Job / Offer (Offer is a reverse one-to-one). ``Max`` aggregates
  fold the latest interview date and the candidate's latest resume-upload datetime
  into the same row without an N+1 loop. ``Max`` is chosen deliberately: it is
  unaffected by the row fan-out that annotating two multi-valued relations
  (``interviews`` and ``candidate__resumes``) otherwise causes.

* **Date filter field (documented)** — the resolved date range is applied to the
  **candidate's most recent resume-upload date** (``Max(candidate__resumes__uploaded_at)``),
  compared in the project timezone. This is the "Resume Upload Date" column, so the
  filter and the displayed date are consistent. Candidates with no resume are
  therefore excluded from a date-filtered report (they have no upload date to match).

Everything here returns ``(bytes, content_type, filename)`` so the API layer can
stream it straight back as an attachment. Generation is a single query plus an
in-memory render, comfortably inside the BRD's 30s budget.
"""
from __future__ import annotations

import csv
import datetime
import io
from collections import OrderedDict

from django.db.models import Count, F, Max
from django.utils import timezone

from .models import (
    Candidate,
    CandidateJobMapping,
    Job,
    Offer,
    RecruitmentStatus,
    ReportConfiguration,
)


# Report types (Module 8). ``candidate`` is the default candidate-data export;
# ``job`` is the per-job report with a status summary; ``openings`` is the
# openings-over-time analytics report (posted / closed / still-open by period).
REPORT_TYPE_CANDIDATE = "candidate"
REPORT_TYPE_JOB = "job"
REPORT_TYPE_OPENINGS = "openings"
VALID_REPORT_TYPES = (REPORT_TYPE_CANDIDATE, REPORT_TYPE_JOB, REPORT_TYPE_OPENINGS)

# Openings-report grains (trailing periods, per the reference periodsFor).
GRAIN_WEEK = "week"
GRAIN_MONTH = "month"
GRAIN_YEAR = "year"
VALID_GRAINS = (GRAIN_WEEK, GRAIN_MONTH, GRAIN_YEAR)


# ---------------------------------------------------------------------------
# Column registry (BRD Module 8 — exact list & order)
# ---------------------------------------------------------------------------
# (key, human header). ``columns`` on a ReportConfiguration is a subset of these
# keys; when provided it selects & reorders, otherwise all 15 are emitted.
COLUMN_REGISTRY: "OrderedDict[str, str]" = OrderedDict(
    [
        ("candidate_name", "Candidate Name"),
        ("job_role", "Job Role"),
        ("recruiter", "Recruiter"),
        ("resume_upload_date", "Resume Upload Date"),
        ("total_experience", "Total Experience"),
        ("relevant_experience", "Relevant Experience"),
        ("current_location", "Current Location"),
        ("current_salary", "Current Salary"),
        ("expected_salary", "Expected Salary"),
        ("notice_period", "Notice Period"),
        ("last_working_day", "Last Working Day"),
        ("candidate_status", "Candidate Status"),
        ("job_status", "Job Status"),
        ("interview_date", "Interview Date"),
        ("offer_status", "Offer Status"),
    ]
)
ALL_COLUMN_KEYS = list(COLUMN_REGISTRY.keys())

# Job report (report_type=job) — one row per Job.
JOB_COLUMN_REGISTRY: "OrderedDict[str, str]" = OrderedDict(
    [
        ("job_id", "Job ID"),
        ("job_role", "Job Role"),
        ("department", "Department"),
        ("location", "Location"),
        ("job_status", "Job Status"),
        ("number_of_openings", "Openings"),
        ("candidate_count", "Candidates"),
    ]
)
JOB_COLUMN_KEYS = list(JOB_COLUMN_REGISTRY.keys())

# Openings report (report_type=openings) — one row per period (reference openingsRows).
OPENINGS_COLUMN_REGISTRY: "OrderedDict[str, str]" = OrderedDict(
    [
        ("period", "Period"),
        ("roles_posted", "Roles Posted"),
        ("openings_posted", "Openings Posted"),
        ("roles_closed", "Roles Closed"),
        ("openings_closed", "Openings Closed"),
        ("roles_open", "Roles Still Open"),
        ("openings_open", "Openings Still Open"),
        ("candidates_joined", "Candidates Joined"),
    ]
)
OPENINGS_COLUMN_KEYS = list(OPENINGS_COLUMN_REGISTRY.keys())


def resolve_columns(columns=None, registry=None):
    """Return the ordered list of (key, header) pairs to emit.

    ``columns`` is an optional ordered subset of the registry keys (e.g. a
    ReportConfiguration.columns list or a ``?columns=`` query value). Unknown keys
    are ignored; an empty / falsy selection defaults to the full registry.
    ``registry`` selects the candidate (default) or job column set.
    """
    registry = registry if registry is not None else COLUMN_REGISTRY
    keys = [k for k in (columns or []) if k in registry]
    if not keys:
        keys = list(registry.keys())
    return [(k, registry[k]) for k in keys]


# ---------------------------------------------------------------------------
# Date-range resolver (BRD Module 8 filters)
# ---------------------------------------------------------------------------
def resolve_date_range(date_filter, custom_start=None, custom_end=None):
    """Map a ReportConfiguration.DateFilter value → ``(start_date, end_date)``.

    Both bounds are inclusive ``datetime.date`` values in the project's local
    timezone (weeks start Monday). ``CUSTOM`` uses the supplied custom bounds.

    Raises ``ValueError`` on an unknown filter or an incomplete/invalid custom range.
    """
    DF = ReportConfiguration.DateFilter
    today = timezone.localdate()

    if date_filter == DF.TODAY:
        return today, today
    if date_filter == DF.YESTERDAY:
        d = today - datetime.timedelta(days=1)
        return d, d
    if date_filter == DF.THIS_WEEK:
        start = today - datetime.timedelta(days=today.weekday())  # Monday
        return start, start + datetime.timedelta(days=6)
    if date_filter == DF.LAST_WEEK:
        this_monday = today - datetime.timedelta(days=today.weekday())
        start = this_monday - datetime.timedelta(days=7)
        return start, start + datetime.timedelta(days=6)
    if date_filter == DF.THIS_MONTH:
        start = today.replace(day=1)
        return start, _end_of_month(start)
    if date_filter == DF.LAST_MONTH:
        first_this = today.replace(day=1)
        last_prev = first_this - datetime.timedelta(days=1)
        return last_prev.replace(day=1), last_prev
    if date_filter == DF.THIS_YEAR:
        return today.replace(month=1, day=1), today.replace(month=12, day=31)
    if date_filter == DF.CUSTOM:
        if not custom_start or not custom_end:
            raise ValueError("Custom date range requires both start and end dates.")
        if custom_end < custom_start:
            raise ValueError("Custom end date must be on or after the start date.")
        return custom_start, custom_end

    raise ValueError(f"Unknown date filter: {date_filter!r}")


def _end_of_month(first_day):
    if first_day.month == 12:
        next_first = first_day.replace(year=first_day.year + 1, month=1, day=1)
    else:
        next_first = first_day.replace(month=first_day.month + 1, day=1)
    return next_first - datetime.timedelta(days=1)


def date_filter_label(date_filter, start, end):
    """Human-readable caption, e.g. 'This Month (2026-07-01 to 2026-07-31)'."""
    label = ReportConfiguration.DateFilter(date_filter).label
    return f"{label} ({start.isoformat()} to {end.isoformat()})"


# ---------------------------------------------------------------------------
# Row builder (single efficient query)
# ---------------------------------------------------------------------------
def _local_bounds(start, end):
    """Inclusive local-date range → aware datetime bounds [start 00:00, end 23:59:59]."""
    tz = timezone.get_current_timezone()
    start_dt = datetime.datetime.combine(start, datetime.time.min)
    end_dt = datetime.datetime.combine(end, datetime.time.max)
    if timezone.is_naive(start_dt):
        start_dt = timezone.make_aware(start_dt, tz)
        end_dt = timezone.make_aware(end_dt, tz)
    return start_dt, end_dt


def _fmt_decimal(value):
    """Decimal → float for numeric-friendly output; None → ''."""
    return float(value) if value is not None else ""


def build_report_rows(date_filter, custom_start=None, custom_end=None, columns=None):
    """Return ``(column_pairs, rows)`` for the resolved filter.

    * ``column_pairs`` — ordered ``[(key, header), ...]`` (see :func:`resolve_columns`).
    * ``rows`` — a list of ``OrderedDict`` keyed by the selected column keys, values
      already coerced to render-friendly primitives (str / float / date / '' ).

    One DB round-trip: ``select_related`` + ``Max`` aggregates, no per-row queries.
    """
    start, end = resolve_date_range(date_filter, custom_start, custom_end)
    start_dt, end_dt = _local_bounds(start, end)
    column_pairs = resolve_columns(columns)

    qs = (
        CandidateJobMapping.objects.select_related("candidate", "job", "offer")
        .annotate(
            latest_interview_date=Max("interviews__interview_date"),
            resume_upload_dt=Max("candidate__resumes__uploaded_at"),
            offer_status_val=F("offer__offer_status"),
        )
        # Filter on the candidate's most recent resume-upload date (documented above).
        .filter(resume_upload_dt__gte=start_dt, resume_upload_dt__lte=end_dt)
        # Exclude mappings whose candidate or job is in the Recycle Bin (this is
        # CandidateJobMapping-grain, so soft-deleted parents don't auto-filter).
        .filter(candidate__deleted_at__isnull=True, job__deleted_at__isnull=True)
        .order_by("candidate__full_name", "job__job_id")
    )

    cand_status_labels = dict(Candidate.Status.choices)
    job_status_labels = dict(Job.Status.choices)
    offer_status_labels = dict(Offer.Status.choices)
    keys = [k for k, _ in column_pairs]

    rows = []
    for m in qs.iterator():
        cand = m.candidate
        job = m.job
        upload_dt = m.resume_upload_dt
        full = {
            "candidate_name": cand.full_name,
            "job_role": job.job_role,
            "recruiter": m.recruiter_name or "",
            "resume_upload_date": (
                timezone.localtime(upload_dt).date() if upload_dt else ""
            ),
            "total_experience": _fmt_decimal(cand.total_experience_years),
            "relevant_experience": _fmt_decimal(cand.relevant_experience_years),
            "current_location": cand.current_location or "",
            "current_salary": _fmt_decimal(cand.current_salary),
            "expected_salary": _fmt_decimal(cand.expected_salary),
            "notice_period": (
                cand.notice_period_days if cand.notice_period_days is not None else ""
            ),
            "last_working_day": cand.last_working_day or "",
            "candidate_status": cand_status_labels.get(
                cand.candidate_status, cand.candidate_status
            ),
            "job_status": job_status_labels.get(job.job_status, job.job_status),
            "interview_date": m.latest_interview_date or "",
            "offer_status": (
                offer_status_labels.get(m.offer_status_val, "")
                if m.offer_status_val
                else ""
            ),
        }
        rows.append(OrderedDict((k, full[k]) for k in keys))

    return column_pairs, rows


# ---------------------------------------------------------------------------
# Job report (report_type=job) — one row per Job + status summary
# ---------------------------------------------------------------------------
def build_job_report_rows(date_filter, custom_start=None, custom_end=None, columns=None):
    """Return ``(column_pairs, rows, summary)`` for the job report.

    * ``rows`` — one ``OrderedDict`` per Job created within the resolved date
      range (filtered on ``Job.created_at``), with a ``candidate_count`` annotation.
    * ``summary`` — ordered ``[(label, count), ...]`` of jobs by status
      (Open / In Progress / Closed / On Hold) plus a Total, over the same range.

    One DB round-trip for the rows plus one aggregate query for the summary.
    """
    start, end = resolve_date_range(date_filter, custom_start, custom_end)
    start_dt, end_dt = _local_bounds(start, end)
    column_pairs = resolve_columns(columns, registry=JOB_COLUMN_REGISTRY)
    keys = [k for k, _ in column_pairs]

    qs = (
        Job.objects.filter(created_at__gte=start_dt, created_at__lte=end_dt)
        .annotate(candidate_count=Count("candidate_mappings", distinct=True))
        .order_by("job_id")
    )

    job_status_labels = dict(Job.Status.choices)
    rows = []
    for job in qs.iterator():
        full = {
            "job_id": job.job_id,
            "job_role": job.job_role,
            "department": job.department,
            "location": job.location,
            "job_status": job_status_labels.get(job.job_status, job.job_status),
            "number_of_openings": job.number_of_openings,
            "candidate_count": job.candidate_count,
        }
        rows.append(OrderedDict((k, full[k]) for k in keys))

    # --- Status summary (single aggregate query over the same date range) ----
    # Build from a *clean* queryset: ``qs`` carries the ``candidate_count`` join to
    # ``candidate_mappings``, which would make ``Count("id")`` count mappings, not
    # jobs (a job with N candidates counted N times). Re-filter Job directly so the
    # status counts are one row per job.
    counts = {
        r["job_status"]: r["n"]
        for r in (
            Job.objects.filter(created_at__gte=start_dt, created_at__lte=end_dt)
            .values("job_status")
            .annotate(n=Count("id"))
        )
    }
    summary = [
        (label, counts.get(value, 0)) for value, label in Job.Status.choices
    ]
    summary.append(("Total", sum(counts.values())))

    return column_pairs, rows, summary


# ---------------------------------------------------------------------------
# Openings report (report_type=openings) — openings over time by period
# ---------------------------------------------------------------------------
def _openings_of(number_of_openings):
    """Openings a job contributes: max(1, number_of_openings) (reference openingsOf)."""
    try:
        return max(1, int(number_of_openings or 0))
    except (TypeError, ValueError):
        return 1


def _month_start(year, month):
    """First-of-month datetime (aware) for the given year/month (1-based month)."""
    return _local_bounds(datetime.date(year, month, 1), datetime.date(year, month, 1))[0]


def periods_for(grain):
    """Trailing reporting periods as ``[(label, from_dt, to_dt), ...]`` (aware).

    Mirrors the reference ``periodsFor``:
      * ``week``  — last 12 weeks (weeks start Monday), label ``"W/c DD/MM/YYYY"``.
      * ``month`` — last 12 months, label ``"Mon YYYY"`` (e.g. ``"Jul 2026"``).
      * ``year``  — last 5 years, label ``"YYYY"``.
    """
    if grain not in VALID_GRAINS:
        raise ValueError(f"Unknown grain: {grain!r}. Use week, month or year.")

    today = timezone.localdate()
    out = []

    if grain == GRAIN_WEEK:
        monday = today - datetime.timedelta(days=today.weekday())
        for i in range(11, -1, -1):
            from_date = monday - datetime.timedelta(days=i * 7)
            to_date = from_date + datetime.timedelta(days=6)
            from_dt, to_dt = _local_bounds(from_date, to_date)
            out.append((f"W/c {from_date.strftime('%d/%m/%Y')}", from_dt, to_dt))
    elif grain == GRAIN_MONTH:
        for i in range(11, -1, -1):
            # Walk back i months from the first of the current month.
            month_index = (today.year * 12 + (today.month - 1)) - i
            year, month0 = divmod(month_index, 12)
            month = month0 + 1
            first = datetime.date(year, month, 1)
            last = _end_of_month(first)
            from_dt, to_dt = _local_bounds(first, last)
            out.append((first.strftime("%b %Y"), from_dt, to_dt))
    else:  # GRAIN_YEAR
        for i in range(4, -1, -1):
            year = today.year - i
            from_dt, to_dt = _local_bounds(
                datetime.date(year, 1, 1), datetime.date(year, 12, 31)
            )
            out.append((str(year), from_dt, to_dt))

    return out


def build_openings_report_rows(grain=GRAIN_MONTH, columns=None):
    """Return ``(column_pairs, rows)`` for the openings report (reference openingsRows).

    For each trailing period compute: Roles/Openings posted (by ``Job.created_at``),
    Roles/Openings closed (by ``Job.closed_at``), Roles/Openings still open (created
    on/before the period end and not yet closed by then), and Candidates joined
    (a status transition to JOINED within the period, from ``RecruitmentStatus``).
    A final ``Total`` row sums the flows (still-open shows the latest period value).

    Efficient: one values() fetch of jobs and one of joined-transition timestamps;
    all bucketing is in memory.
    """
    periods = periods_for(grain)
    column_pairs = resolve_columns(columns, registry=OPENINGS_COLUMN_REGISTRY)
    keys = [k for k, _ in column_pairs]

    # ``Job.objects`` already excludes soft-deleted jobs (SoftDeleteManager);
    # exclude joined-transitions of soft-deleted candidates for the same reason.
    jobs = list(Job.objects.values("created_at", "closed_at", "number_of_openings"))
    joined_dts = list(
        RecruitmentStatus.objects.filter(
            new_status=Candidate.Status.JOINED,
            candidate__deleted_at__isnull=True,
        ).values_list("changed_at", flat=True)
    )

    def _in(dt, frm, to):
        return dt is not None and frm <= dt <= to

    full_rows = []
    for label, frm, to in periods:
        posted = [j for j in jobs if _in(j["created_at"], frm, to)]
        closed = [j for j in jobs if _in(j["closed_at"], frm, to)]
        open_at_end = [
            j
            for j in jobs
            if j["created_at"] is not None
            and j["created_at"] <= to
            and (j["closed_at"] is None or j["closed_at"] > to)
        ]
        joined = [d for d in joined_dts if _in(d, frm, to)]

        full_rows.append(
            {
                "period": label,
                "roles_posted": len(posted),
                "openings_posted": sum(_openings_of(j["number_of_openings"]) for j in posted),
                "roles_closed": len(closed),
                "openings_closed": sum(_openings_of(j["number_of_openings"]) for j in closed),
                "roles_open": len(open_at_end),
                "openings_open": sum(_openings_of(j["number_of_openings"]) for j in open_at_end),
                "candidates_joined": len(joined),
            }
        )

    # Totals row: sum the flow columns; still-open reflects the latest period.
    if full_rows:
        last = full_rows[-1]
        totals = {
            "period": "Total",
            "roles_posted": sum(r["roles_posted"] for r in full_rows),
            "openings_posted": sum(r["openings_posted"] for r in full_rows),
            "roles_closed": sum(r["roles_closed"] for r in full_rows),
            "openings_closed": sum(r["openings_closed"] for r in full_rows),
            "roles_open": last["roles_open"],
            "openings_open": last["openings_open"],
            "candidates_joined": sum(r["candidates_joined"] for r in full_rows),
        }
        full_rows.append(totals)

    rows = [OrderedDict((k, r[k]) for k in keys) for r in full_rows]
    return column_pairs, rows


# ---------------------------------------------------------------------------
# Rendering helpers
# ---------------------------------------------------------------------------
def _cell_text(value):
    """Uniform string rendering for CSV/PDF cells."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float):
        # Trim trailing .0 for whole numbers (salaries/experience read cleaner).
        return str(int(value)) if value.is_integer() else str(value)
    return str(value)


def _timestamp():
    return timezone.localtime().strftime("%Y-%m-%d %H:%M")


def _filename(ext):
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    return f"recruitment_report_{stamp}.{ext}"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def generate_excel(column_pairs, rows, caption="", summary=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = [h for _, h in column_pairs]
    ncols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Recruitment Report"

    # Title row (filter label + generated timestamp), merged across all columns.
    title = f"Recruitment Report — {caption}" if caption else "Recruitment Report"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"Generated: {_timestamp()}").font = Font(
        italic=True, size=9, color="666666"
    )
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    header_row = 3
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(rows, start=header_row + 1):
        for col, (key, _) in enumerate(column_pairs, start=1):
            value = row[key]
            ws.cell(row=r, column=col, value=value if value != "" else None)

    # Optional status summary block below the table (job report).
    if summary:
        summary_row = header_row + 1 + len(rows) + 1
        ws.cell(row=summary_row, column=1, value="Summary — Jobs by Status").font = Font(
            bold=True, size=11
        )
        for i, (label, value) in enumerate(summary, start=summary_row + 1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=(label == "Total"))
            ws.cell(row=i, column=2, value=value).font = Font(bold=(label == "Total"))

    # Freeze the header (everything above the first data row stays visible).
    ws.freeze_panes = f"A{header_row + 1}"

    # Auto-ish column widths from header + data string lengths (capped).
    for col, (key, header) in enumerate(column_pairs, start=1):
        width = len(header)
        for row in rows:
            width = max(width, len(_cell_text(row[key])))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    return (
        buffer.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        _filename("xlsx"),
    )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def generate_csv(column_pairs, rows, caption="", summary=None):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    if caption:
        writer.writerow([f"Recruitment Report — {caption}"])
        writer.writerow([f"Generated: {_timestamp()}"])
        writer.writerow([])
    writer.writerow([h for _, h in column_pairs])
    for row in rows:
        writer.writerow([_cell_text(row[key]) for key, _ in column_pairs])

    # Optional status summary block below the table (job report).
    if summary:
        writer.writerow([])
        writer.writerow(["Summary — Jobs by Status"])
        for label, value in summary:
            writer.writerow([label, value])

    # utf-8-sig so Excel opens the CSV with correct encoding.
    data = buffer.getvalue().encode("utf-8-sig")
    return data, "text/csv", _filename("csv")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def generate_pdf(column_pairs, rows, caption="", summary=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    styles = getSampleStyleSheet()
    header_style = styles["Heading1"]
    caption_style = styles["Normal"]
    cell_style = styles["BodyText"]
    cell_style.fontSize = 7
    cell_style.leading = 9

    headers = [Paragraph(f"<b>{h}</b>", cell_style) for _, h in column_pairs]
    table_data = [headers]
    for row in rows:
        table_data.append(
            [Paragraph(_cell_text(row[key]) or "", cell_style) for key, _ in column_pairs]
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Recruitment Report",
    )

    story = [Paragraph("Recruitment Report", header_style)]
    if caption:
        story.append(Paragraph(caption, caption_style))
    story.append(Paragraph(f"Generated: {_timestamp()}", caption_style))
    story.append(Spacer(1, 6 * mm))

    if rows:
        table = Table(table_data, repeatRows=1)  # header repeats on every page
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0B0B0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5FA")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        story.append(table)
    else:
        story.append(Paragraph("No records match the selected filter.", caption_style))

    # Optional status summary block below the table (job report).
    if summary:
        story.append(Spacer(1, 8 * mm))
        story.append(Paragraph("Summary — Jobs by Status", styles["Heading2"]))
        summary_data = [
            [Paragraph(f"<b>{label}</b>" if label == "Total" else label, cell_style),
             Paragraph(f"<b>{value}</b>" if label == "Total" else str(value), cell_style)]
            for label, value in summary
        ]
        summary_table = Table(summary_data, colWidths=[60 * mm, 30 * mm])
        summary_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0B0B0")),
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F2F5FA")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        story.append(summary_table)

    doc.build(story)
    return buffer.getvalue(), "application/pdf", _filename("pdf")


# ---------------------------------------------------------------------------
# Format dispatch
# ---------------------------------------------------------------------------
_GENERATORS = {
    ReportConfiguration.ExportFormat.EXCEL: generate_excel,
    ReportConfiguration.ExportFormat.CSV: generate_csv,
    ReportConfiguration.ExportFormat.PDF: generate_pdf,
}


def generate_report(
    export_format,
    date_filter,
    custom_start=None,
    custom_end=None,
    columns=None,
    report_type=REPORT_TYPE_CANDIDATE,
    grain=GRAIN_MONTH,
):
    """End-to-end: build rows then render → ``(bytes, content_type, filename)``.

    ``export_format`` must be a ReportConfiguration.ExportFormat value
    (``EXCEL`` / ``CSV`` / ``PDF``). ``report_type`` selects the candidate export
    (default), the per-job report (with a status summary), or the openings report
    (openings over time; uses ``grain`` — ``week`` | ``month`` | ``year`` — and
    ignores the date filter). Raises ``ValueError`` on a bad format, unknown report
    type/grain, or invalid date range.
    """
    generator = _GENERATORS.get(export_format)
    if generator is None:
        raise ValueError(f"Unsupported export format: {export_format!r}")
    if report_type not in VALID_REPORT_TYPES:
        raise ValueError(
            f"Unknown report_type: {report_type!r}. Use 'candidate', 'job' or 'openings'."
        )

    # The openings report uses fixed trailing periods (by grain), not a date filter.
    if report_type == REPORT_TYPE_OPENINGS:
        grain_labels = {
            GRAIN_WEEK: "Last 12 weeks",
            GRAIN_MONTH: "Last 12 months",
            GRAIN_YEAR: "Last 5 years",
        }
        caption = f"Openings over time — {grain_labels.get(grain, grain)}"
        column_pairs, rows = build_openings_report_rows(grain, columns)
        return generator(column_pairs, rows, caption=caption)

    start, end = resolve_date_range(date_filter, custom_start, custom_end)
    caption = date_filter_label(date_filter, start, end)

    if report_type == REPORT_TYPE_JOB:
        column_pairs, rows, summary = build_job_report_rows(
            date_filter, custom_start, custom_end, columns
        )
        return generator(column_pairs, rows, caption=caption, summary=summary)

    column_pairs, rows = build_report_rows(date_filter, custom_start, custom_end, columns)
    return generator(column_pairs, rows, caption=caption)
